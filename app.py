import streamlit as st
import pandas as pd
import json
import os
import base64
import requests
from google.cloud import bigquery
from google.oauth2 import service_account
from datetime import datetime, timedelta
from pytz import timezone
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================
# 0. CONFIGURACIÓN
# ============================================================
st.set_page_config(
    page_title="Centro de Monitoreo - amb", 
    page_icon="🌧️", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================
# 1. CONFIGURACIÓN Y LOGO
# ============================================================
logo_path = os.path.join(os.path.dirname(__file__), "amb_4_punto_cero.jpg")
try:
    st.sidebar.image(logo_path, use_column_width=True)
except:
    st.sidebar.warning("Logo no cargado")

st.title("🌧️ Centro de Monitoreo: Red Meteorológica amb")
colombia_tz = timezone('America/Bogota')
st.caption(f"🕐 Última actualización: {datetime.now(colombia_tz).strftime('%Y-%m-%d %H:%M:%S')} (hora Colombia)")

# ============================================================
# 2. CONFIGURACIÓN
# ============================================================
AUTOR = "Mauricio Mora"
VERSION = "2.0"
SISTEMA = "Sistema Automatizado de Monitoreo"

# ============================================================
# 3. CONFIGURACIÓN DEL AGENTE IA
# ============================================================
AGENTE_API_URL = "https://querybigqueryamb-ia-661926446380.us-central1.run.app"

# ============================================================
# 4. UMBRALES Y ALERTAS (SEMÁFORO)
# ============================================================
umbrales = {
    "El_Pajal": {"amarilla": 12.3, "naranja": 15.1, "roja": 20.4},
    "Yerbabuena": {"amarilla": 10.9, "naranja": 20.0, "roja": 40.8},
    "La_Mariana": {"amarilla": 11.7, "naranja": 18.0, "roja": 35.0},
    "Vegas_del_Quemado": {"amarilla": 27.2, "naranja": 36.8, "roja": 55.8}
}

# Nivel de rebose del embalse
NIVEL_REBOSE_EMBALSE = 885.75

def obtener_alerta(precipitacion, estacion):
    if estacion == "Monsalve": 
        return "AZUL", "🛠️ En Aprendizaje", "#3399FF", "0s"
    if estacion == "Embalse":
        return "EMBALSE", "🌊 Nivel de Embalse", "#00BFFF", "0s"
    if estacion not in umbrales: 
        return "GRIS", "☁️ Sin umbrales definidos", "#CCCCCC", "0s"
    
    u = umbrales[estacion]
    if precipitacion >= u["roja"]: 
        return "ROJA", f"🚨 ROJA: Excede {u['roja']}mm", "#FF4B4B", "0.5s"
    elif precipitacion >= u["naranja"]: 
        return "NARANJA", f"⚠️ NARANJA: Excede {u['naranja']}mm", "#FF9933", "1s"
    elif precipitacion >= u["amarilla"]: 
        return "AMARILLA", f"🟡 AMARILLA: Excede {u['amarilla']}mm", "#FFFF00", "2s"
    elif precipitacion > 0: 
        return "VERDE", "✅ Lluvia Normal", "#00CC96", "0s"
    return "GRIS", "☁️ Sin lluvia", "#CCCCCC", "0s"

def evaluar_nivel_embalse(nivel_actual):
    excedente = nivel_actual - NIVEL_REBOSE_EMBALSE
    if excedente >= 0:
        return "🔴 EXCEDENTE", "#FF4B4B", f"{excedente:.2f} msnm por encima del nivel de rebose", excedente
    else:
        return "🟢 NORMAL", "#00CC96", f"{abs(excedente):.2f} msnm por debajo del nivel de rebose", excedente

# ============================================================
# 5. CLIENTE BIGQUERY
# ============================================================
@st.cache_resource
def init_bigquery_client():
    try:
        json_str = base64.b64decode(st.secrets["GCP_JSON_B64"]).decode('utf-8')
        key_dict = json.loads(json_str)
        creds = service_account.Credentials.from_service_account_info(key_dict)
        return bigquery.Client(credentials=creds, project=key_dict["project_id"])
    except Exception as e:
        st.error(f"❌ Error al conectar con BigQuery: {e}")
        st.stop()

client = init_bigquery_client()

# ============================================================
# 6. FUNCIONES DE DATOS
# ============================================================
@st.cache_data(ttl=300)
def get_last_reading(estacion):
    try:
        query = f"""
        SELECT * FROM `gen-lang-client-0342049346.amb_hidrologia.telemetria_estaciones` 
        WHERE id_estacion = '{estacion}' 
        ORDER BY SAFE_CAST(timestamp AS TIMESTAMP) DESC 
        LIMIT 1
        """
        df = client.query(query).to_dataframe()
        if not df.empty:
            df['timestamp'] = pd.to_datetime(df['timestamp']).dt.tz_localize('UTC').dt.tz_convert('America/Bogota')
        return df
    except Exception as e:
        st.error(f"❌ Error al obtener última lectura: {e}")
        return pd.DataFrame()

@st.cache_data(ttl=600)
def get_historical_data_range(estacion, fecha_inicio, fecha_fin):
    try:
        if isinstance(fecha_inicio, datetime):
            fecha_inicio_str = fecha_inicio.strftime('%Y-%m-%d')
        else:
            fecha_inicio_str = fecha_inicio
        
        if isinstance(fecha_fin, datetime):
            fecha_fin_str = fecha_fin.strftime('%Y-%m-%d')
        else:
            fecha_fin_str = fecha_fin
        
        query = f"""
        SELECT * FROM `gen-lang-client-0342049346.amb_hidrologia.telemetria_estaciones` 
        WHERE id_estacion = '{estacion}' 
        AND SAFE_CAST(timestamp AS TIMESTAMP) >= TIMESTAMP('{fecha_inicio_str} 00:00:00')
        AND SAFE_CAST(timestamp AS TIMESTAMP) <= TIMESTAMP('{fecha_fin_str} 23:59:59')
        ORDER BY SAFE_CAST(timestamp AS TIMESTAMP) DESC 
        LIMIT 50000
        """
        
        df = client.query(query).to_dataframe()
        if not df.empty:
            df['timestamp'] = pd.to_datetime(df['timestamp']).dt.tz_localize('UTC').dt.tz_convert('America/Bogota')
        return df
    except Exception as e:
        st.error(f"❌ Error en datos históricos: {str(e)}")
        return pd.DataFrame()

# ============================================================
# 7. FUNCIONES DE EXPORTACIÓN
# ============================================================
def preparar_df_para_exportar(df):
    df_export = df.copy()
    if 'timestamp' in df_export.columns:
        df_export['timestamp'] = df_export['timestamp'].dt.tz_localize(None)
    return df_export

def generar_excel_con_formato(df, nombre_estacion, periodo_descripcion):
    df_export = preparar_df_para_exportar(df)
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name='Datos', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Datos']
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, len(df_export.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column].width = adjusted_width
        
        metadata = pd.DataFrame({
            'Propiedad': ['Sistema', 'Estación', 'Período', 'Fecha de exportación', 'Total de registros', 'Versión'],
            'Valor': [
                SISTEMA,
                nombre_estacion,
                periodo_descripcion,
                datetime.now(colombia_tz).strftime('%Y-%m-%d %H:%M:%S'),
                len(df_export),
                VERSION
            ]
        })
        metadata.to_excel(writer, sheet_name='Metadatos', index=False)
        
        metadata_sheet = writer.sheets['Metadatos']
        for col in range(1, 3):
            cell = metadata_sheet.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    return output.getvalue()

def generar_resumen_estadistico(df):
    if df.empty:
        return "No hay datos disponibles"
    
    resumen = []
    resumen.append("📊 RESUMEN ESTADÍSTICO")
    resumen.append("=" * 40)
    resumen.append("")
    resumen.append("📋 Datos generados automáticamente por el sistema")
    resumen.append("")
    
    columnas_numericas = ['temperatura', 'precipitacion', 'humedad', 'presion', 'velocidad_viento', 'voltaje_bateria']
    for col in columnas_numericas:
        if col in df.columns:
            datos = df[col].dropna()
            if not datos.empty:
                resumen.append(f"📈 {col.upper()}:")
                resumen.append(f"   • Promedio: {datos.mean():.2f}")
                resumen.append(f"   • Máximo:  {datos.max():.2f}")
                resumen.append(f"   • Mínimo:  {datos.min():.2f}")
                resumen.append(f"   • Registros: {len(datos)}")
                resumen.append("")
    
    return "\n".join(resumen)

# ============================================================
# 8. FUNCIÓN PARA CONSULTAR AGENTE IA
# ============================================================
def consultar_agente_ia(pregunta):
    try:
        headers = {"Content-Type": "application/json"}
        payload = {"prompt": pregunta}
        
        response = requests.post(AGENTE_API_URL, json=payload, headers=headers, timeout=30)
        
        if response.status_code == 200:
            data = response.json()
            
            if isinstance(data, dict):
                if "mensaje" in data:
                    return {
                        "status": "ok",
                        "mensaje": data["mensaje"],
                        "datos": data.get("datos", []),
                        "estacion": data.get("estacion", ""),
                        "variable": data.get("variable", "precipitacion"),
                        "contexto": data.get("contexto", {}),
                        "fuente": data.get("fuente", ""),
                        "raw": data
                    }
                elif data.get("status") == "ok":
                    if "mensaje" in data and data["mensaje"]:
                        return {
                            "status": "ok",
                            "mensaje": data["mensaje"],
                            "datos": data.get("datos", []),
                            "estacion": data.get("estacion", ""),
                            "variable": data.get("variable", "precipitacion"),
                            "contexto": data.get("contexto", {}),
                            "fuente": data.get("fuente", ""),
                            "raw": data
                        }
                    else:
                        return formatear_respuesta_agente(data)
                else:
                    return {
                        "status": "error",
                        "mensaje": data.get("mensaje", "Error al procesar la consulta.")
                    }
            else:
                return {
                    "status": "error",
                    "mensaje": "Respuesta del agente en formato inesperado"
                }
        else:
            return {
                "status": "error",
                "mensaje": f"Error en la API: {response.status_code}"
            }
    except requests.exceptions.Timeout:
        return {
            "status": "error",
            "mensaje": "⏱️ Tiempo de espera agotado."
        }
    except requests.exceptions.ConnectionError:
        return {
            "status": "error",
            "mensaje": "❌ No se pudo conectar al agente IA."
        }
    except Exception as e:
        return {
            "status": "error",
            "mensaje": f"❌ Error: {str(e)}"
        }

def formatear_respuesta_agente(data):
    estacion = data.get("estacion", "Estación")
    variable = data.get("variable")
    
    if variable is None:
        variable = "precipitacion"
    
    datos = data.get("datos", [])
    contexto = data.get("contexto", {})
    fuente = data.get("fuente", "")
    periodo = data.get("periodo", {})
    mensaje = data.get("mensaje", "")
    
    if mensaje:
        return {
            "status": "ok",
            "mensaje": mensaje,
            "datos": datos,
            "estacion": estacion,
            "raw": data
        }
    
    nombres_variables = {
        "precipitacion": "Precipitación",
        "temperatura": "Temperatura",
        "humedad": "Humedad",
        "caudal": "Caudal",
        "nivel": "Nivel",
        "presion": "Presión",
        "velocidad_viento": "Velocidad del Viento"
    }
    nombre_variable = nombres_variables.get(variable, variable.capitalize())
    
    mensaje = f"🌧️ **{estacion}** - {nombre_variable}\n\n"
    
    if periodo:
        fecha_inicio = periodo.get("fecha_inicio", "")
        fecha_fin = periodo.get("fecha_fin", "")
        if fecha_inicio and fecha_fin:
            if fecha_inicio == fecha_fin:
                mensaje += f"📅 **Fecha:** {fecha_inicio}\n\n"
            else:
                mensaje += f"📅 **Período:** {fecha_inicio} al {fecha_fin}\n\n"
    
    if datos:
        valores = [d.get("valor", 0) for d in datos]
        promedio = sum(valores) / len(valores) if valores else 0
        maximo = max(valores) if valores else 0
        minimo = min(valores) if valores else 0
        
        mensaje += f"📊 **Estadísticas:**\n"
        mensaje += f"• Promedio: {promedio:.2f}\n"
        mensaje += f"• Máximo: {maximo:.2f}\n"
        mensaje += f"• Mínimo: {minimo:.2f}\n"
        mensaje += f"• Registros: {len(datos)}\n\n"
        
        mensaje += f"📈 **Datos por fecha:**\n"
        for d in datos[:10]:
            fecha = d.get("fecha", "")
            valor = d.get("valor", 0)
            mensaje += f"• {fecha}: {valor:.2f}\n"
        if len(datos) > 10:
            mensaje += f"\n*... y {len(datos) - 10} registros más*"
    else:
        mensaje += "📊 No se encontraron datos para el período consultado."
    
    if contexto:
        mensaje += "\n\n📍 **Contexto geográfico:**\n"
        if contexto.get("cuenca"):
            mensaje += f"• Cuenca: {contexto['cuenca']}\n"
        if contexto.get("afecta"):
            mensaje += f"• Afecta: {contexto['afecta']}\n"
    
    if fuente:
        mensaje += f"\n📡 **Fuente:** {fuente}"
    
    return {
        "status": "ok",
        "mensaje": mensaje,
        "datos": datos,
        "estacion": estacion,
        "raw": data
    }

def verificar_agente_ia():
    try:
        response = requests.post(
            AGENTE_API_URL,
            json={"prompt": "ping"},
            timeout=5
        )
        return response.status_code == 200 or response.status_code == 400
    except:
        return False

# ============================================================
# 9. FUNCIONES DE VISUALIZACIÓN
# ============================================================
def create_embalse_chart(df_hist):
    try:
        if df_hist.empty:
            return None
        
        df_ordenado = df_hist.sort_values('timestamp')
        fig = go.Figure()
        
        fig.add_trace(go.Scatter(
            x=df_ordenado['timestamp'],
            y=df_ordenado['temperatura'],
            mode='lines',
            name='Nivel del embalse',
            line=dict(color='#00BFFF', width=2),
            fill='tozeroy',
            fillcolor='rgba(0, 191, 255, 0.2)'
        ))
        
        fig.add_hline(
            y=NIVEL_REBOSE_EMBALSE,
            line_dash="dash",
            line_color="red",
            line_width=2,
            annotation_text=f"Nivel de rebose: {NIVEL_REBOSE_EMBALSE} msnm",
            annotation_position="top right"
        )
        
        fig.update_layout(
            title='Nivel del Embalse (msnm)',
            xaxis_title='Fecha/Hora',
            yaxis_title='Nivel (msnm)',
            height=400,
            template='plotly_white',
            hovermode='x unified'
        )
        return fig
    except Exception as e:
        st.warning(f"No se pudo generar el gráfico del embalse: {e}")
        return None

# ============================================================
# 9.5 FUNCIONES PARA EDV (EXTENSÓMETROS)
# ============================================================

@st.cache_data(ttl=600)
def get_edv_data(extensometro='izquierdo'):
    """
    Obtiene los datos del EDV desde BigQuery
    """
    try:
        table = f"edv_{extensometro}"
        query = f"""
        SELECT 
            fecha,
            anillo,
            lectura,
            cota_referencia,
            cota,
            asiento,
            dist_datum,
            notas
        FROM `gen-lang-client-0342049346.amb_hidrologia.{table}`
        ORDER BY fecha DESC, CAST(anillo AS INT64) DESC
        """
        df = client.query(query).to_dataframe()
        if not df.empty:
            df['fecha'] = pd.to_datetime(df['fecha']).dt.tz_localize('UTC').dt.tz_convert('America/Bogota')
        return df
    except Exception as e:
        st.error(f"❌ Error al obtener datos EDV: {e}")
        return pd.DataFrame()

def create_edv_profile(df, fecha_seleccionada=None, titulo="Perfil de Deformaciones"):
    """
    Crea el perfil de deformaciones para una fecha específica
    """
    if fecha_seleccionada is None:
        fecha_seleccionada = df['fecha'].max()
    
    df_fecha = df[df['fecha'].dt.date == fecha_seleccionada.date()]
    
    if df_fecha.empty:
        return None
    
    df_fecha = df_fecha.sort_values('anillo', ascending=False)
    
    fig = go.Figure()
    
    fig.add_trace(go.Scatter(
        x=df_fecha['asiento'],
        y=df_fecha['anillo'],
        mode='lines+markers',
        name=f'Perfil {fecha_seleccionada.strftime("%d/%m/%Y")}',
        line=dict(color='#FF4B4B', width=3),
        marker=dict(size=12, color='#FF4B4B')
    ))
    
    fig.add_vline(x=0, line_dash="dash", line_color="gray", line_width=1)
    fig.add_hline(y=0, line_dash="dash", line_color="green", line_width=2, 
                  annotation_text="FONDO", annotation_position="bottom right")
    
    fig.update_layout(
        title=f'{titulo} - {fecha_seleccionada.strftime("%d/%m/%Y")}',
        xaxis_title='Asiento (cm)',
        yaxis_title='Anillo',
        template='plotly_white',
        height=500,
        hovermode='y unified'
    )
    return fig

def create_edv_multiple_profiles(df, max_profiles=10):
    fechas_unicas = sorted(df['fecha'].unique(), reverse=True)
    if len(fechas_unicas) > max_profiles:
        step = len(fechas_unicas) // max_profiles
        fechas_seleccionadas = fechas_unicas[::step][:max_profiles]
        if fechas_unicas[0] not in fechas_seleccionadas:
            fechas_seleccionadas[0] = fechas_unicas[0]
    else:
        fechas_seleccionadas = fechas_unicas
    
    fig = go.Figure()
    colores = px.colors.sequential.Reds[::-1] + px.colors.sequential.Blues[::-1]
    
    for i, fecha in enumerate(fechas_seleccionadas):
        df_fecha = df[df['fecha'].dt.date == fecha.date()]
        if df_fecha.empty:
            continue
        
        df_fecha = df_fecha.sort_values('anillo', ascending=False)
        color_idx = i % len(colores)
        color = colores[color_idx]
        width = 2 + (len(fechas_seleccionadas) - i) * 0.2
        width = min(width, 4)
        
        fig.add_trace(go.Scatter(
            x=df_fecha['asiento'],
            y=df_fecha['anillo'],
            mode='lines+markers',
            name=fecha.strftime('%d/%m/%Y'),
            line=dict(color=color, width=width),
            marker=dict(size=6, color=color)
        ))
    
    fig.add_vline(x=0, line_dash="dash", line_color="gray", line_width=1)
    fig.add_hline(y=0, line_dash="dash", line_color="green", line_width=2, 
                  annotation_text="FONDO", annotation_position="bottom right")
    
    fig.update_layout(
        title='Evolución de Deformaciones',
        xaxis_title='Asiento (cm)',
        yaxis_title='Anillo',
        template='plotly_white',
        height=600,
        hovermode='y unified',
        legend=dict(x=1.02, y=1, bgcolor='rgba(255,255,255,0.9)')
    )
    return fig

def create_edv_timeline(df, anillos_seleccionados=None):
    if anillos_seleccionados is None:
        anillos_disponibles = sorted(df['anillo'].unique(), key=lambda x: int(x))
        if len(anillos_disponibles) >= 3:
            anillos_seleccionados = [
                anillos_disponibles[-1],
                anillos_disponibles[len(anillos_disponibles)//2],
                anillos_disponibles[0]
            ]
        else:
            anillos_seleccionados = anillos_disponibles
    
    fig = go.Figure()
    colores = ['#FF4B4B', '#FF9933', '#00CC96', '#3399FF', '#FF69B4']
    
    for i, anillo in enumerate(anillos_seleccionados):
        df_anillo = df[df['anillo'] == anillo].sort_values('fecha')
        if not df_anillo.empty:
            fig.add_trace(go.Scatter(
                x=df_anillo['fecha'],
                y=df_anillo['asiento'],
                mode='lines+markers',
                name=f'Anillo {anillo}',
                line=dict(color=colores[i % len(colores)], width=2),
                marker=dict(size=6)
            ))
    
    fig.update_layout(
        title='Evolución del Asiento por Anillo',
        xaxis_title='Fecha',
        yaxis_title='Asiento (cm)',
        template='plotly_white',
        height=400,
        hovermode='x unified'
    )
    return fig

def mostrar_seccion_edv():
    st.markdown("---")
    st.subheader("📏 Instrumentación Geotécnica - Extensómetros (EDV)")
    st.caption("Mediciones de deformación vertical del terreno alrededor del embalse")
    
    col1, col2 = st.columns([1, 2])
    with col1:
        extensometro_seleccionado = st.selectbox(
            "Seleccione Extensómetro:",
            ["izquierdo", "derecho"],
            index=0
        )
    with col2:
        tipo_vista = st.selectbox(
            "Tipo de vista:",
            ["Perfil Individual", "Evolución Histórica", "Evolución Temporal"],
            index=0
        )
    
    with st.spinner("🔄 Cargando datos del extensómetro..."):
        df_edv = get_edv_data(extensometro_seleccionado)
    
    if df_edv.empty:
        st.warning("⚠️ No hay datos disponibles para este extensómetro")
        return
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("📊 Total registros", f"{len(df_edv):,}")
    with col2:
        st.metric("🔢 Anillos", df_edv['anillo'].nunique())
    with col3:
        st.metric("📅 Primera lectura", df_edv['fecha'].min().strftime('%d/%m/%Y'))
    with col4:
        st.metric("📅 Última lectura", df_edv['fecha'].max().strftime('%d/%m/%Y'))
    
    if tipo_vista == "Perfil Individual":
        fechas_disponibles = sorted(df_edv['fecha'].unique(), reverse=True)
        fecha_seleccionada = st.selectbox(
            "Seleccione fecha para el perfil:",
            fechas_disponibles,
            index=0,
            format_func=lambda x: x.strftime('%d/%m/%Y')
        )
        
        fig = create_edv_profile(df_edv, fecha_seleccionada, f"EDV {extensometro_seleccionado.capitalize()}")
        if fig:
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No hay datos para la fecha seleccionada")
    
    elif tipo_vista == "Evolución Histórica":
        fig = create_edv_multiple_profiles(df_edv, max_profiles=10)
        if fig:
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No hay suficientes datos para mostrar")
    
    elif tipo_vista == "Evolución Temporal":
        anillos_disponibles = sorted(df_edv['anillo'].unique(), key=lambda x: int(x))
        anillos_seleccionados = st.multiselect(
            "Seleccione anillos para visualizar:",
            anillos_disponibles,
            default=[anillos_disponibles[-1], anillos_disponibles[len(anillos_disponibles)//2], anillos_disponibles[0]]
        )
        
        if anillos_seleccionados:
            fig = create_edv_timeline(df_edv, anillos_seleccionados)
            if fig:
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No hay datos para los anillos seleccionados")
        else:
            st.info("Seleccione al menos un anillo para visualizar")
    
    with st.expander("📋 Ver datos detallados"):
        st.dataframe(
            df_edv[['fecha', 'anillo', 'lectura', 'cota', 'asiento', 'dist_datum']],
            use_container_width=True,
            column_config={
                "fecha": st.column_config.DateColumn("Fecha"),
                "anillo": st.column_config.TextColumn("Anillo"),
                "lectura": st.column_config.NumberColumn("Lectura (mm)", format="%.2f"),
                "cota": st.column_config.NumberColumn("Cota (msnm)", format="%.2f"),
                "asiento": st.column_config.NumberColumn("Asiento (cm)", format="%.2f"),
                "dist_datum": st.column_config.NumberColumn("Dist. Datum (m)", format="%.2f"),
            }
        )
        
        csv = df_edv.to_csv(index=False).encode('utf-8-sig')
        st.download_button(
            "📥 Descargar datos (CSV)",
            csv,
            f"edv_{extensometro_seleccionado}.csv",
            "text/csv",
            use_container_width=True
        )

# ============================================================
# 10. INTERFAZ PRINCIPAL
# ============================================================
estaciones = ["La_Mariana", "Yerbabuena", "Vegas_del_Quemado", "El_Pajal", "Monsalve", "Embalse"]
seleccion = st.sidebar.selectbox("Seleccione Estación:", estaciones)

periodos_grafico = {
    "Últimas 24 horas": 24,
    "Últimos 3 días": 72,
    "Últimos 7 días": 168,
    "Últimos 15 días": 360,
    "Último mes": 720
}

if seleccion == "Embalse":
    st.sidebar.markdown("### 📊 Opciones de Histórico")
    periodo_seleccionado = st.sidebar.selectbox(
        "Seleccione período:",
        list(periodos_grafico.keys())
    )
    horas = periodos_grafico[periodo_seleccionado]
else:
    horas = st.sidebar.slider("⏱️ Horas históricas:", 1, 168, 24, step=1)

with st.spinner("🔄 Cargando datos..."):
    df = get_last_reading(seleccion)
    fecha_fin = datetime.now(colombia_tz)
    fecha_inicio = fecha_fin - timedelta(hours=horas)
    df_hist = get_historical_data_range(seleccion, fecha_inicio, fecha_fin)

# ============================================================
# TABS
# ============================================================
tab1, tab2, tab3 = st.tabs(["📊 Situación Actual", "📈 Históricos", "🤖 Asistente IA"])

# ============================================================
# TAB 1: SITUACIÓN ACTUAL
# ============================================================
with tab1:
    if not df.empty:
        row = df.iloc[0]
        st.subheader(f"📡 Real-time: {seleccion}")
        
        if seleccion == "Embalse":
            nivel_actual = float(row.get('temperatura', 0))
            estado, color, mensaje, excedente = evaluar_nivel_embalse(nivel_actual)
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("🌊 Nivel actual", f"{nivel_actual:.2f} msnm", delta=f"{excedente:.2f} msnm")
            with col2:
                st.metric("📏 Nivel de Rebose", f"{NIVEL_REBOSE_EMBALSE:.2f} msnm")
            with col3:
                st.metric("📊 Excédente", f"{excedente:+.2f} msnm")
            
            if excedente >= 0:
                st.error(f"🔴 {estado} - {mensaje}")
                st.warning("⚠️ El embalse está por encima del nivel de rebose. ¡Monitorear constantemente!")
            else:
                st.success(f"🟢 {estado} - {mensaje}")
            
            if 'voltaje_bateria' in row:
                st.metric("🔋 Voltaje", f"{float(row['voltaje_bateria']):.1f} V")
            
            st.info(f"📅 Última lectura: {row['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}")
            
            # Mostrar sección EDV
            mostrar_seccion_edv()
            
        else:
            nombre, msg, color, vel = obtener_alerta(float(row.get('precipitacion', 0)), seleccion)
            st.markdown(f'''
            <div style="background-color:{color}; padding:20px; border-radius:15px; text-align:center; color:black; animation: blink {vel} infinite; border: 2px solid #333;">
                <h2>🚦 {nombre}</h2>
                <b>{msg}</b>
            </div>
            <style>
            @keyframes blink {{
                0%{{opacity:1}} 
                50%{{opacity:0.3}} 
                100%{{opacity:1}}
            }}
            </style>
            ''', unsafe_allow_html=True)
            st.write("")
            
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("🌡️ Temp", f"{float(row['temperatura']):.1f} °C")
            c2.metric("🌧️ Precip", f"{float(row['precipitacion']):.1f} mm")
            c3.metric("💧 Humedad", f"{float(row['humedad']):.1f} %")
            c4.metric("🔋 Voltaje", f"{float(row['voltaje_bateria']):.1f} V")
            
            st.info(f"📅 Última lectura: {row['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}")
            
            if not df_hist.empty:
                st.markdown("### 📊 Estadísticas del Período")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("🔽 Temp Mínima", f"{df_hist['temperatura'].min():.1f}°C")
                with col2:
                    st.metric("🔼 Temp Máxima", f"{df_hist['temperatura'].max():.1f}°C")
                with col3:
                    st.metric("📊 Temp Promedio", f"{df_hist['temperatura'].mean():.1f}°C")
    else:
        st.warning("⚠️ Sin datos.")

# ============================================================
# TAB 2: HISTÓRICOS CON GRÁFICOS Y DESCARGA
# ============================================================
with tab2:
    st.subheader("📈 Series de Tiempo")
    
    if not df_hist.empty:
        if seleccion == "Embalse":
            # ============================================
            # EMBALSE - SOLO GRÁFICO DE NIVEL
            # ============================================
            st.markdown("### 🌊 Nivel del Embalse")
            
            fig_embalse = create_embalse_chart(df_hist)
            if fig_embalse:
                st.plotly_chart(fig_embalse, use_container_width=True)
            
            ultima_lectura = df_hist.iloc[0]
            st.info(f"📊 Último nivel registrado: {ultima_lectura['temperatura']:.2f} msnm")
            
            with st.expander("📋 Ver datos detallados"):
                columnas_embalse = ['timestamp', 'temperatura', 'voltaje_bateria']
                df_embalse_mostrar = df_hist[columnas_embalse].copy()
                df_embalse_mostrar.columns = ['Fecha/Hora', 'Nivel (msnm)', 'Voltaje (V)']
                st.dataframe(df_embalse_mostrar.head(20), use_container_width=True)
            
        else:
            # ============================================
            # OTRAS ESTACIONES - TODOS LOS GRÁFICOS
            # ============================================
            st.markdown("### 🌡️ Temperatura")
            fig_temp = px.line(
                df_hist.sort_values('timestamp'), 
                x='timestamp', 
                y='temperatura',
                title=f'Temperatura - {seleccion}',
                labels={'temperatura': '°C', 'timestamp': 'Fecha/Hora'}
            )
            fig_temp.update_layout(height=300, template='plotly_white', hovermode='x unified')
            if len(df_hist) > 1:
                fig_temp.add_hline(
                    y=df_hist['temperatura'].mean(), 
                    line_dash="dash", 
                    line_color="red",
                    annotation_text=f"Promedio: {df_hist['temperatura'].mean():.1f}°C"
                )
            st.plotly_chart(fig_temp, use_container_width=True)
            
            st.markdown("### 🌧️ Precipitación")
            fig_precip = px.bar(
                df_hist.sort_values('timestamp'), 
                x='timestamp', 
                y='precipitacion',
                title=f'Precipitación - {seleccion}',
                labels={'precipitacion': 'mm', 'timestamp': 'Fecha/Hora'},
                color='precipitacion',
                color_continuous_scale='Blues'
            )
            fig_precip.update_layout(height=300, template='plotly_white')
            st.plotly_chart(fig_precip, use_container_width=True)
            
            if 'humedad' in df_hist.columns:
                st.markdown("### 💧 Humedad")
                fig_humedad = px.line(
                    df_hist.sort_values('timestamp'), 
                    x='timestamp', 
                    y='humedad',
                    title=f'Humedad - {seleccion}',
                    labels={'humedad': '%', 'timestamp': 'Fecha/Hora'}
                )
                fig_humedad.update_layout(height=300, template='plotly_white', hovermode='x unified')
                if len(df_hist) > 1:
                    fig_humedad.add_hline(
                        y=df_hist['humedad'].mean(), 
                        line_dash="dash", 
                        line_color="red",
                        annotation_text=f"Promedio: {df_hist['humedad'].mean():.1f}%"
                    )
                st.plotly_chart(fig_humedad, use_container_width=True)
            
            if 'direccion_del_viento' in df_hist.columns and 'velocidad_viento' in df_hist.columns:
                st.markdown("### 🧭 Rosa de los Vientos")
                df_viento = df_hist.dropna(subset=['direccion_del_viento', 'velocidad_viento'])
                if not df_viento.empty:
                    fig_viento = px.bar_polar(
                        df_viento,
                        r="velocidad_viento",
                        theta="direccion_del_viento",
                        color="velocidad_viento",
                        color_continuous_scale='Viridis',
                        title=f'Rosa de Vientos - {seleccion}',
                        template='plotly_white'
                    )
                    fig_viento.update_layout(height=400)
                    st.plotly_chart(fig_viento, use_container_width=True)
                else:
                    st.info("ℹ️ No hay datos de viento disponibles para esta estación")
    else:
        st.info("ℹ️ No hay datos históricos disponibles para este período")
    
    # ============================================================
    # DESCARGA PERSONALIZADA DE DATOS
    # ============================================================
    st.markdown("---")
    st.subheader("📥 Descarga Personalizada de Datos")
    
    st.markdown("### 📅 Selecciona el período para descargar")
    
    col_periodo1, col_periodo2 = st.columns(2)
    
    with col_periodo1:
        opcion_periodo = st.radio(
            "Período:",
            ["Diario", "Semanal", "Mensual", "Semestral", "Anual"],
            index=0
        )
    
    with col_periodo2:
        opcion_personalizado = st.checkbox("📅 Personalizar fechas")
    
    hoy = datetime.now(colombia_tz)
    
    if opcion_personalizado:
        st.markdown("### 📅 Selecciona las fechas personalizadas")
        col_fecha1, col_fecha2 = st.columns(2)
        with col_fecha1:
            fecha_inicio_descarga = st.date_input(
                "Fecha de inicio:",
                value=hoy - timedelta(days=30),
                max_value=hoy
            )
        with col_fecha2:
            fecha_fin_descarga = st.date_input(
                "Fecha de fin:",
                value=hoy,
                max_value=hoy
            )
        
        if fecha_inicio_descarga > fecha_fin_descarga:
            st.error("❌ La fecha de inicio no puede ser mayor que la fecha de fin")
            st.stop()
        
        periodo_descripcion = f"Personalizado ({fecha_inicio_descarga.strftime('%d/%m/%Y')} - {fecha_fin_descarga.strftime('%d/%m/%Y')})"
    else:
        if opcion_periodo == "Diario":
            fecha_inicio_descarga = hoy - timedelta(days=1)
            fecha_fin_descarga = hoy
            periodo_descripcion = f"Diario ({fecha_inicio_descarga.strftime('%d/%m/%Y')})"
        elif opcion_periodo == "Semanal":
            fecha_inicio_descarga = hoy - timedelta(days=7)
            fecha_fin_descarga = hoy
            periodo_descripcion = f"Semanal ({fecha_inicio_descarga.strftime('%d/%m/%Y')} - {fecha_fin_descarga.strftime('%d/%m/%Y')})"
        elif opcion_periodo == "Mensual":
            fecha_inicio_descarga = hoy - timedelta(days=30)
            fecha_fin_descarga = hoy
            periodo_descripcion = f"Mensual ({fecha_inicio_descarga.strftime('%d/%m/%Y')} - {fecha_fin_descarga.strftime('%d/%m/%Y')})"
        elif opcion_periodo == "Semestral":
            fecha_inicio_descarga = hoy - timedelta(days=180)
            fecha_fin_descarga = hoy
            periodo_descripcion = f"Semestral ({fecha_inicio_descarga.strftime('%d/%m/%Y')} - {fecha_fin_descarga.strftime('%d/%m/%Y')})"
        else:
            fecha_inicio_descarga = hoy - timedelta(days=365)
            fecha_fin_descarga = hoy
            periodo_descripcion = f"Anual ({fecha_inicio_descarga.strftime('%d/%m/%Y')} - {fecha_fin_descarga.strftime('%d/%m/%Y')})"
    
    st.info(f"📊 **Período seleccionado:** {periodo_descripcion}")
    
    if st.button("📥 Cargar datos para este período", use_container_width=True):
        with st.spinner("🔄 Cargando datos históricos..."):
            df_descarga = get_historical_data_range(
                seleccion, 
                fecha_inicio_descarga, 
                fecha_fin_descarga
            )
            
            if not df_descarga.empty:
                st.session_state['df_descarga'] = df_descarga
                st.session_state['periodo_descarga'] = periodo_descripcion
                st.success(f"✅ Datos cargados: {len(df_descarga)} registros")
            else:
                st.warning("⚠️ No hay datos para el período seleccionado")
    
    if 'df_descarga' in st.session_state:
        df_descarga = st.session_state['df_descarga']
        periodo_descarga = st.session_state['periodo_descarga']
        
        with st.expander("📊 Ver resumen estadístico"):
            st.text(generar_resumen_estadistico(df_descarga))
        
        with st.expander("📋 Ver datos cargados"):
            st.dataframe(df_descarga, use_container_width=True)
        
        st.markdown("### 📥 Descargar hoja de datos")
        st.caption("Selecciona el formato para descargar los datos históricos")
        
        col_export1, col_export2 = st.columns(2)
        
        df_export = preparar_df_para_exportar(df_descarga)
        csv_data = df_export.to_csv(index=False).encode('utf-8-sig')
        
        with col_export1:
            try:
                excel_data = generar_excel_con_formato(df_descarga, seleccion, periodo_descarga)
                st.download_button(
                    "📊 Hoja de cálculo (.xlsx)",
                    excel_data,
                    f"{seleccion}_{datetime.now(colombia_tz).strftime('%Y%m%d_%H%M')}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.caption("✅ Formato Excel con metadatos y formato profesional")
            except Exception as e:
                st.error(f"Error al generar Excel: {e}")
                st.download_button(
                    "📊 Hoja de datos (alternativo)",
                    csv_data,
                    f"{seleccion}_{datetime.now(colombia_tz).strftime('%Y%m%d_%H%M')}.csv",
                    "text/csv",
                    use_container_width=True
                )
        
        with col_export2:
            st.download_button(
                "📝 Google Sheets",
                csv_data,
                f"{seleccion}_{datetime.now(colombia_tz).strftime('%Y%m%d_%H%M')}.csv",
                "text/csv",
                use_container_width=True,
                help="Formato CSV compatible con Google Sheets"
            )
            st.caption("📤 Abre en Google Sheets")
        
        st.caption(f"📋 Datos exportados desde el {SISTEMA}")

# ============================================================
# TAB 3: ASISTENTE IA
# ============================================================
with tab3:
    st.subheader("🤖 Asistente IA - Centro de Monitoreo")
    st.markdown("Pregunta sobre niveles, caudales, lluvias y estado de las estaciones.")
    
    with st.spinner("🔌 Verificando conexión..."):
        if verificar_agente_ia():
            st.success("✅ Agente IA conectado")
        else:
            st.warning("⚠️ No se pudo conectar al agente IA. Verifica la configuración.")
    
    if "messages" not in st.session_state:
        st.session_state.messages = []
    
    chat_container = st.container()
    
    with chat_container:
        for message in st.session_state.messages:
            with st.chat_message(message["role"]):
                st.markdown(message["content"])
    
    prompt = st.chat_input("Escribe tu pregunta sobre las estaciones...")
    
    if prompt:
        st.session_state.messages.append({"role": "user", "content": prompt})
        
        with st.chat_message("user"):
            st.markdown(prompt)
        
        with st.chat_message("assistant"):
            with st.spinner("🤔 Analizando tu pregunta..."):
                resultado = consultar_agente_ia(prompt)
                
                if resultado.get("status") == "ok":
                    respuesta = resultado.get("mensaje", "✅ Consulta procesada exitosamente.")
                    st.markdown(respuesta)
                    st.session_state.messages.append({"role": "assistant", "content": respuesta})
                elif resultado.get("status") == "sin_datos":
                    mensaje = f"ℹ️ {resultado.get('mensaje', 'No se encontraron datos.')}"
                    st.info(mensaje)
                    st.session_state.messages.append({"role": "assistant", "content": mensaje})
                else:
                    error_msg = resultado.get("mensaje", "❌ Error al procesar la consulta.")
                    st.error(error_msg)
                    st.session_state.messages.append({"role": "assistant", "content": error_msg})
        
        st.rerun()
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("🗑️ Limpiar conversación", use_container_width=True):
            st.session_state.messages = []
            st.rerun()
    with col2:
        with st.expander("💡 Ejemplos de preguntas"):
            st.markdown("""
            **🌊 Sobre el embalse:**
            - ¿Cómo está el nivel del embalse?
            - Nivel a las 3 de la mañana
            - Promedio nivel últimas 24 horas
            
            **🌧️ Sobre estaciones:**
            - ¿Cuánto llovió en El_Pajal ayer?
            - Temperatura máxima en La_Mariana este mes
            - Temperatura a las 15:00 en El_Pajal
            
            **📊 Consultas avanzadas:**
            - Comparar lluvias entre El_Pajal y Yerbabuena
            - ¿Qué relación hay entre la lluvia y el nivel del embalse?
            """)

# ============================================================
# 11. FOOTER Y SIDEBAR
# ============================================================
st.sidebar.markdown("---")
st.sidebar.markdown("**Desarrollado por Mauricio Mora**")
st.sidebar.caption("📊 Datos actualizados cada 5 minutos")

with st.sidebar.expander("🌊 Información del Embalse"):
    st.write(f"**Nivel de Rebose:** {NIVEL_REBOSE_EMBALSE} msnm")
    if not df.empty and seleccion == "Embalse":
        nivel_actual = float(df.iloc[0].get('temperatura', 0))
        excedente = nivel_actual - NIVEL_REBOSE_EMBALSE
        st.write(f"**Nivel Actual:** {nivel_actual:.2f} msnm")
        if excedente >= 0:
            st.error(f"**Excédente:** +{excedente:.2f} msnm")
        else:
            st.success(f"**Déficit:** {excedente:.2f} msnm")

with st.sidebar.expander("🤖 Estado del Agente IA"):
    st.write(f"**URL:** {AGENTE_API_URL}")
    st.write("**Status:** ✅ Activo")
    st.write("**Capacidades:**")
    st.write("- 📊 Consultas SCADA (2026+)")
    st.write("- 📜 Históricos (2004-2025)")
    st.write("- 🌊 Análisis de embalse")
    st.write("- ⏰ Consultas por hora específica")

# ============================================================
# 12. INFORMACIÓN DE EDV EN SIDEBAR
# ============================================================
with st.sidebar.expander("📏 Extensómetros (EDV)"):
    st.write("**Datos cargados en BigQuery:**")
    st.write("- EDV Izquierdo: 4,577 registros")
    st.write("- EDV Derecho: 4,323 registros")
    st.write("**Período:** 2013-2025")
    st.write("**Estado:** ✅ Activo")
    
    st.markdown("---")
    st.markdown("### 📱 Registrar nueva medición")
    st.markdown("""
    **📊 Usa la hoja de cálculo para registrar:**
    
    [📝 Abrir hoja de EDV](https://docs.google.com/spreadsheets/d/TU_ID_AQUI)
    
    **Instrucciones:**
    1. Carga lecturas anteriores (menú EDV)
    2. Ingresa lecturas actuales
    3. Genera formato impresión
    4. Sube a BigQuery
    
    🔒 **Seguridad:**
    - Solo administrador puede subir datos
    - Operadores pueden llenar y firmar
    """)
